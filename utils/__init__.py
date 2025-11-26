"""Shared utility functions for routes."""
from flask import request, jsonify
from flask_login import login_required, current_user
from functools import wraps
from datetime import datetime
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

from models import VolunteerRecord


def require_user_type(user_type):
    """Decorator to require specific user type"""
    def decorator(f):
        @login_required
        def wrapper(*args, **kwargs):
            if not current_user.is_authenticated or current_user.user_type != user_type:
                if request.is_json:
                    return jsonify({'error': 'Not authenticated'}), 401
                from flask import redirect, url_for
                return redirect(url_for('auth.login'))
            return f(*args, **kwargs)
        wrapper.__name__ = f.__name__
        return wrapper
    return decorator


def generate_excel_from_records(records, filename_prefix="volunteer_records", user_display_name=None):
    """Helper function to generate Excel file from VolunteerRecord list."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Volunteer Records"
    
    # Header row
    headers = ['Project Name', 'Category', 'Organization', 'Date', 'Certified Hours', 'Points Earned', 'Status']
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Data rows
    for row_idx, record in enumerate(records, start=2):
        project = record.project
        organization = project.organization if project else None
        
        ws.cell(row=row_idx, column=1, value=project.title if project else 'Unknown Project')
        ws.cell(row=row_idx, column=2, value=project.category if project else 'N/A')
        ws.cell(row=row_idx, column=3, value=organization.display_name or organization.username if organization else 'Unknown Organization')
        ws.cell(row=row_idx, column=4, value=record.completed_at.strftime('%Y-%m-%d') if record.completed_at else 'N/A')
        ws.cell(row=row_idx, column=5, value=record.hours)
        ws.cell(row=row_idx, column=6, value=record.points)
        
        # Status mapping
        status_display = {
            'approved': 'Certified',
            'pending': 'Pending',
            'rejected': 'Rejected'
        }
        ws.cell(row=row_idx, column=7, value=status_display.get(record.status, record.status))
    
    # Auto-adjust column widths
    column_widths = [30, 15, 25, 12, 15, 15, 12]
    for col_idx, width in enumerate(column_widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Generate filename with user display_name at the front
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    if user_display_name:
        # Sanitize display_name for filename (remove invalid characters)
        safe_name = "".join(c for c in user_display_name if c.isalnum() or c in (' ', '-', '_')).strip()
        safe_name = safe_name.replace(' ', '_')
        filename = f"{safe_name}_{filename_prefix}_{timestamp}.xlsx"
    else:
        filename = f"{filename_prefix}_{timestamp}.xlsx"
    
    return output, filename








