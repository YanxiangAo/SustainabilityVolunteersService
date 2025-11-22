from flask import Blueprint, render_template, request, redirect, url_for, session, jsonify, current_app, send_file, make_response
from flask_login import login_user, logout_user, login_required, current_user
from sqlalchemy import or_
from datetime import datetime, timedelta
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

from models import db, User, Project, Registration, VolunteerRecord, Badge, UserBadge, SystemSettings, Comment
from forms import parse_login_form, parse_register_form

bp = Blueprint('main', __name__)

# Helper function to check user type
def require_user_type(user_type):
    """Decorator to require specific user type"""
    def decorator(f):
        @login_required
        def wrapper(*args, **kwargs):
            if not current_user.is_authenticated or current_user.user_type != user_type:
                if request.is_json:
                    return jsonify({'error': 'Not authenticated'}), 401
                return redirect(url_for('main.login'))
            return f(*args, **kwargs)
        wrapper.__name__ = f.__name__
        return wrapper
    return decorator

# Routes
@bp.route('/')
def index():
    return render_template('index.html')

@bp.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        form = parse_login_form(request)
        user = User.query.filter(
            User.user_type == form.user_type,
            or_(User.username == form.username_or_email, User.email == form.username_or_email)
        ).first()

        if user and user.check_password(form.password):
            # Check if user is active
            if hasattr(user, 'is_active') and not user.is_active:
                return render_template('login.html', error='Your account has been disabled. Please contact an administrator.')
            
            # Use Flask-Login to log in the user
            login_user(user, remember=form.remember)

            # Ensure we don't persist the session unless the user requested it
            session.permanent = False
            if not form.remember:
                session['_remember'] = 'clear'
                session.pop('_remember_seconds', None)
            
            # Keep session data for backward compatibility
            session['user_type'] = user.user_type
            session['username'] = user.username
            
            if user.user_type == 'participant':
                return redirect(url_for('main.participant_dashboard'))
            elif user.user_type == 'organization':
                return redirect(url_for('main.organization_dashboard'))
            elif user.user_type == 'admin':
                return redirect(url_for('main.admin_panel'))
        else:
            return render_template('login.html', error='Invalid credentials')
    
    return render_template('login.html')

@bp.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        form = parse_register_form(request)
        current_app.logger.info(f"Register POST received: user_type={form.user_type}, username={form.username}, email={form.email}")

        # Validate user_type
        if form.user_type not in ('participant', 'organization'):
            return render_template('login.html', error='Invalid user type')
        
        # Check if user already exists
        if User.query.filter_by(username=form.username).first():
            return render_template('login.html', error='Username already exists')
        
        if User.query.filter_by(email=form.email).first():
            return render_template('login.html', error='Email already exists')
        
        # Create new user
        try:
            user = User()
            user.set_password(form.password)
            db.session.add(user)
            db.session.commit()
            current_app.logger.info(f"Register success: id={user.id}, username={user.username}, type={user.user_type}")
        except Exception as e:
            current_app.logger.exception("Register failed")
            db.session.rollback()
            return render_template('login.html', error=f'Register failed: {str(e)}')
        
        return redirect(url_for('main.login'))
    
    return render_template('login.html')

@bp.route('/logout')
@login_required
def logout():
    logout_user()
    # Ensure remember-me cookies/session flags are cleared
    session.pop('_remember', None)
    session.pop('_remember_seconds', None)
    session.clear()
    
    response = make_response(redirect(url_for('main.index')))
    remember_cookie_name = current_app.config.get('REMEMBER_COOKIE_NAME', 'remember_token')
    response.delete_cookie(remember_cookie_name)
    return response

@bp.route('/participant/dashboard')
@login_required
def participant_dashboard():
    if current_user.user_type != 'participant':
        return redirect(url_for('main.login'))
    
    return render_template('participant_dashboard.html', user=current_user)

@bp.route('/organization/dashboard')
@login_required
def organization_dashboard():
    if current_user.user_type != 'organization':
        return redirect(url_for('main.login'))
    
    return render_template('organization_dashboard.html', user=current_user)

@bp.route('/admin/panel')
@login_required
def admin_panel():
    if current_user.user_type != 'admin':
        return redirect(url_for('main.login'))
    
    return render_template('admin_panel.html', user=current_user)

@bp.route('/project/<int:project_id>')
def project_detail(project_id):
    project = Project.query.get_or_404(project_id)
    # Get organization info
    organization = User.query.get(project.organization_id)
    # Get registration count (only active registrations)
    active_statuses = ('registered', 'approved')
    registration_count = Registration.query.filter(
        Registration.project_id == project.id,
        Registration.status.in_(active_statuses)
    ).count()
    # Get comments for display
    comments = Comment.query.filter_by(project_id=project.id).order_by(Comment.created_at.desc()).limit(20).all()
    comments_data = []
    for comment in comments:
        comments_data.append({
            'id': comment.id,
            'user_name': comment.user.display_name or comment.user.username if comment.user else 'Unknown',
            'user_type': comment.user.user_type.title() if comment.user else 'Unknown',
            'comment': comment.content,
            'created_at': comment.created_at.strftime('%Y-%m-%d %H:%M') if comment.created_at else None
        })
    
    # Check if current user is registered or is the organization owner
    is_registered = False
    can_comment = False
    user_type = None
    if current_user.is_authenticated:
        user_type = current_user.user_type
        # Check if user is registered (for participants)
        if user_type == 'participant':
            existing_reg = Registration.query.filter(
                Registration.user_id == current_user.id,
                Registration.project_id == project.id,
                Registration.status.in_(active_statuses + ('completed',))
            ).first()
            is_registered = existing_reg is not None
            can_comment = is_registered
        # Organization owner can always comment on their own projects
        elif user_type == 'organization':
            can_comment = (project.organization_id == current_user.id)
    
    return render_template('project_detail.html', 
                         project=project, 
                         organization=organization,
                         registration_count=registration_count,
                         comments=comments_data,
                         is_registered=is_registered,
                         can_comment=can_comment,
                         user_type=user_type)

@bp.route('/demo/project')
def demo_project_detail():
    """Demo route showing project detail page with seeded database data."""
    project = Project.query.filter_by(status='approved').order_by(Project.date.asc()).first()
    if not project:
        return redirect(url_for('main.index'))

    organization = User.query.get(project.organization_id)
    registration_count = Registration.query.filter(
        Registration.project_id == project.id,
        Registration.status != 'cancelled'
    ).count()
    registrations = Registration.query.filter_by(project_id=project.id)\
        .order_by(Registration.created_at.desc())\
        .limit(5)\
        .all()

    comments = []
    for registration in registrations:
        participant = registration.user
        comments.append({
            'user_name': participant.display_name or participant.username,
            'user_type': participant.user_type.title(),
            'comment': f"Looking forward to {project.title}!",
            'created_at': registration.created_at.strftime('%Y-%m-%d %H:%M'),
            'reply': None
        })

    return render_template(
        'project_detail.html',
        project=project,
        organization=organization,
        registration_count=registration_count,
        comments=comments,
        is_demo=True
    )

@bp.route('/volunteer-record')
@login_required
def volunteer_record():
    if current_user.user_type != 'participant':
        return redirect(url_for('main.login'))
    
    records = VolunteerRecord.query.filter_by(user_id=current_user.id).order_by(VolunteerRecord.completed_at.desc()).all()
    
    # Calculate statistics
    total_hours = sum(r.hours for r in records if r.status == 'approved')
    total_points = sum(r.points for r in records if r.status == 'approved')
    completed_count = len([r for r in records if r.status == 'approved'])
    
    # Prepare records data with project and organization info
    records_data = []
    years_set = set()
    categories_set = set()
    
    for record in records:
        project = record.project
        organization = project.organization if project else None
        
        # Collect years and categories for filters
        if record.completed_at:
            years_set.add(record.completed_at.year)
        if project and project.category:
            categories_set.add(project.category)
        
        records_data.append({
            'record': {
                'id': record.id,
                'hours': record.hours,
                'points': record.points,
                'status': record.status,
                'completed_at': record.completed_at.strftime('%Y-%m-%d') if record.completed_at else None
            },
            'project': {
                'id': project.id if project else None,
                'title': project.title if project else 'Unknown Project',
                'category': project.category if project else None
            } if project else None,
            'organization': {
                'display_name': organization.display_name if organization else None,
                'username': organization.username if organization else None
            } if organization else None
        })
    
    return render_template('volunteer_record.html', 
                         user=current_user, 
                         records_data=records_data,
                         total_hours=total_hours,
                         total_points=total_points,
                         completed_count=completed_count,
                         available_years=sorted(years_set, reverse=True),
                         available_categories=sorted(categories_set))


def _generate_excel_from_records(records, filename_prefix="volunteer_records", user_display_name=None):
    """Helper function to generate Excel file from VolunteerRecord list."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Volunteer Records"
    
    # Header row
    headers = ['Project Name', 'Category', 'Organization', 'Date', 'Certified Hours', 'Points Earned', 'Status']
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Data rows
    for row_idx, record in enumerate(records, start=2):
        project = record.project
        organization = project.organization if project else None
        
        ws.cell(row=row_idx, column=1, value=project.title if project else 'Unknown Project')
        ws.cell(row=row_idx, column=2, value=project.category if project else 'N/A')
        ws.cell(row=row_idx, column=3, value=organization.display_name or organization.username if organization else 'Unknown Organization')
        ws.cell(row=row_idx, column=4, value=record.completed_at.strftime('%Y-%m-%d') if record.completed_at else 'N/A')
        ws.cell(row=row_idx, column=5, value=record.hours)
        ws.cell(row=row_idx, column=6, value=record.points)
        
        # Status mapping
        status_display = {
            'approved': 'Certified',
            'pending': 'Pending',
            'rejected': 'Rejected'
        }
        ws.cell(row=row_idx, column=7, value=status_display.get(record.status, record.status))
    
    # Auto-adjust column widths
    column_widths = [30, 15, 25, 12, 15, 15, 12]
    for col_idx, width in enumerate(column_widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Generate filename with user display_name at the front
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    if user_display_name:
        # Sanitize display_name for filename (remove invalid characters)
        safe_name = "".join(c for c in user_display_name if c.isalnum() or c in (' ', '-', '_')).strip()
        safe_name = safe_name.replace(' ', '_')
        filename = f"{safe_name}_{filename_prefix}_{timestamp}.xlsx"
    else:
        filename = f"{filename_prefix}_{timestamp}.xlsx"
    
    return output, filename


@bp.route('/api/participant/export-all-records')
@login_required
def api_export_all_records():
    """Export all volunteer records for the current participant."""
    if current_user.user_type != 'participant':
        return jsonify({'error': 'Not authenticated'}), 401
    
    # Get all records
    records = VolunteerRecord.query.filter_by(user_id=current_user.id).order_by(VolunteerRecord.completed_at.desc()).all()
    
    # Get user display_name (fallback to username if not set)
    user_display_name = current_user.display_name or current_user.username
    
    output, filename = _generate_excel_from_records(records, "all_volunteer_records", user_display_name)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


@bp.route('/api/participant/export-filtered-records', methods=['POST'])
@login_required
def api_export_filtered_records():
    """Export filtered volunteer records for the current participant."""
    if current_user.user_type != 'participant':
        return jsonify({'error': 'Not authenticated'}), 401
    
    data = request.get_json()
    year_filter = data.get('year')
    category_filter = data.get('category')
    
    # Get all records
    records = VolunteerRecord.query.filter_by(user_id=current_user.id).order_by(VolunteerRecord.completed_at.desc()).all()
    
    # Apply filters
    filtered_records = []
    for record in records:
        # Year filter
        if year_filter:
            if not record.completed_at or record.completed_at.year != int(year_filter):
                continue
        
        # Category filter
        if category_filter:
            if not record.project or record.project.category != category_filter:
                continue
        
        filtered_records.append(record)
    
    # Get user display_name (fallback to username if not set)
    user_display_name = current_user.display_name or current_user.username
    
    output, filename = _generate_excel_from_records(filtered_records, "filtered_volunteer_records", user_display_name)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

# ============================================================================
# RESTful API Routes
# ============================================================================

# Projects Resource
@bp.route('/api/v1/projects', methods=['GET'])
def api_projects_list():
    """Get all approved projects (for homepage, exclude expired projects)."""
    # Support query parameters
    status = request.args.get('status', 'approved')
    available = request.args.get('available', 'false').lower() == 'true'
    today = datetime.utcnow().date()
    
    query = Project.query
    if status:
        query = query.filter_by(status=status)
    if available:
        query = query.filter(Project.date >= today)
    
    projects = query.order_by(Project.date.asc()).all()
    
    result = []
    for p in projects:
        current_participants = sum(1 for r in p.registrations if r.status != 'cancelled')
        project_data = {
            'id': p.id,
            'title': p.title,
            'category': p.category,
            'date': p.date.strftime('%Y-%m-%d') if p.date else None,
            'location': p.location,
            'rating': p.rating,
            'max_participants': p.max_participants,
            'current_participants': current_participants,
            'status': p.status,
            'organization': {
                'id': p.organization_id,
                'name': p.organization.display_name or p.organization.username if p.organization else None
            }
        }
        # Only include if not expired when available=true
        if not available or (p.date and p.date >= today):
            result.append(project_data)
    
    return jsonify(result)

@bp.route('/api/v1/projects/<int:project_id>', methods=['GET'])
def api_project_detail(project_id):
    """Get a single project by ID."""
    project = Project.query.get_or_404(project_id)
    organization = project.organization
    
    return jsonify({
        'id': project.id,
        'title': project.title,
        'description': project.description,
        'category': project.category,
        'date': project.date.strftime('%Y-%m-%d') if project.date else None,
        'location': project.location,
        'rating': project.rating,
        'max_participants': project.max_participants,
        'current_participants': sum(1 for r in project.registrations if r.status != 'cancelled'),
        'duration': project.duration,
        'points': project.points,
        'status': project.status,
        'requirements': project.requirements,
        'organization': {
            'id': organization.id if organization else None,
            'name': organization.display_name or organization.username if organization else None,
            'email': organization.email if organization else None
        } if organization else None
    })

@bp.route('/api/v1/projects', methods=['POST'])
@login_required
def api_projects_create():
    """Create a new project."""
    if current_user.user_type != 'organization':
        return jsonify({'error': 'Only organizations can create projects'}), 403
    
    data = request.form if request.form else request.get_json()
    
    project = Project(
        title=data.get('title'),
        description=data.get('description'),
        category=data.get('category'),
        organization_id=current_user.id,
        date=datetime.strptime(data.get('date'), '%Y-%m-%d').date() if data.get('date') else None,
        location=data.get('location'),
        max_participants=int(data.get('max_participants', 0)),
        duration=float(data.get('duration', 0)),
        points=int(data.get('points', 0)),
        status='pending',
        requirements=data.get('requirements', '')
    )
    db.session.add(project)
    db.session.commit()
    
    return jsonify({
        'id': project.id,
        'title': project.title,
        'status': project.status,
        'message': 'Project created successfully'
    }), 201

@bp.route('/api/v1/projects/<int:project_id>', methods=['PATCH'])
@login_required
def api_projects_update(project_id):
    """Update a project (status, etc.)."""
    project = Project.query.get_or_404(project_id)
    data = request.get_json() or {}
    
    # Check permissions
    if current_user.user_type == 'admin':
        # Admin can update status
        if 'status' in data:
            project.status = data['status']
    elif current_user.user_type == 'organization' and project.organization_id == current_user.id:
        # Organization can update their own projects (but not status to approved/rejected)
        if 'status' in data and data['status'] in ('approved', 'rejected'):
            return jsonify({'error': 'Cannot change status to approved/rejected'}), 403
        # Allow other updates
        for key in ['title', 'description', 'category', 'location', 'max_participants', 'duration', 'points', 'requirements']:
            if key in data:
                setattr(project, key, data[key])
    else:
        return jsonify({'error': 'Unauthorized'}), 403
    
    db.session.commit()
    return jsonify({
        'id': project.id,
        'status': project.status,
        'message': 'Project updated successfully'
    })

# Dashboard Resource
@bp.route('/api/v1/users/me/dashboard', methods=['GET'])
@login_required
def api_users_me_dashboard():
    """Get current user's dashboard data."""
    user_type = current_user.user_type
    
    if user_type == 'participant':
        # Get registrations for the user
        user_registrations = Registration.query.filter_by(user_id=current_user.id).order_by(Registration.created_at.desc()).all()
        registration_payload = []
        for registration in user_registrations:
            project = registration.project
            status_label = registration.status.replace('_', ' ').title()
            progress = 0
            if registration.status == 'completed':
                progress = 100
            elif registration.status == 'in_progress':
                progress = 40
            elif registration.status == 'approved':
                progress = 60

            registration_payload.append({
                'id': project.id,
                'title': project.title,
                'organization_name': project.organization.display_name or project.organization.username if project.organization else None,
                'date': project.date.strftime('%Y-%m-%d'),
                'status': status_label,
                'progress': progress
            })

        # Badge data
        all_badges = Badge.query.order_by(Badge.id.asc()).all()
        user_badges = {ub.badge_id: ub for ub in UserBadge.query.filter_by(user_id=current_user.id).all()}

        badges_payload = []
        for badge in all_badges:
            user_badge = user_badges.get(badge.id)
            badges_payload.append({
                'code': badge.code,
                'name': badge.name,
                'description': badge.description,
                'earned': bool(user_badge and user_badge.earned),
                'accent_color': badge.accent_color,
                'background_color': badge.background_color,
                'icon': badge.icon
            })

        # Calculate statistics
        approved_records = VolunteerRecord.query.filter_by(user_id=current_user.id, status='approved').all()
        total_hours = sum(r.hours for r in approved_records)
        total_points = sum(r.points for r in approved_records)
        completed_count = len(approved_records)
        upcoming_count = Registration.query.join(Project).filter(
            Registration.user_id == current_user.id,
            Registration.status.in_(('registered', 'approved')),
            Project.date >= datetime.utcnow().date()
        ).count()

        return jsonify({
            'user': {
                'display_name': current_user.display_name or current_user.username
            },
            'statistics': {
                'total_hours': total_hours,
                'total_points': total_points,
                'completed': completed_count,
                'upcoming': upcoming_count
            },
            'registrations': registration_payload,
            'badges': badges_payload
        })
    
    elif user_type == 'organization':
        # Get organization's projects
        projects = Project.query.filter_by(organization_id=current_user.id).all()
        
        active_projects = sum(1 for p in projects if p.status == 'approved')
        active_registration_statuses = ('registered', 'approved')
        total_participants = sum(
            Registration.query.filter(
                Registration.project_id == p.id,
                Registration.status.in_(active_registration_statuses)
            ).count()
            for p in projects
        )
        completed_projects = sum(1 for p in projects if p.status == 'completed')
        pending_projects = sum(1 for p in projects if p.status == 'pending')
        
        projects_payload = []
        for project in projects:
            registrations = Registration.query.filter_by(project_id=project.id).all()
            projects_payload.append({
                'id': project.id,
                'title': project.title,
                'status': project.status,
                'date': project.date.strftime('%Y-%m-%d') if project.date else None,
                'location': project.location,
                'max_participants': project.max_participants,
                'current_participants': sum(1 for r in registrations if r.status in active_registration_statuses),
                'rating': project.rating
            })
        
        # Get recent projects from the last week
        week_ago = datetime.utcnow() - timedelta(days=7)
        recent_projects = Project.query.filter(
            Project.organization_id == current_user.id,
            Project.created_at >= week_ago,
            Project.status == 'approved'
        ).order_by(Project.created_at.desc()).limit(8).all()
        
        recent_projects_payload = []
        for project in recent_projects:
            registrations = Registration.query.filter_by(project_id=project.id).all()
            recent_projects_payload.append({
                'id': project.id,
                'title': project.title,
                'status': project.status,
                'date': project.date.strftime('%Y-%m-%d') if project.date else None,
                'location': project.location,
                'created_at': project.created_at.strftime('%Y-%m-%d') if project.created_at else None,
                'current_participants': sum(1 for r in registrations if r.status in active_registration_statuses),
                'max_participants': project.max_participants,
                'rating': project.rating,
                'organization_name': project.organization.display_name or project.organization.username if project.organization else None
            })
        
        return jsonify({
            'statistics': {
                'active_projects': active_projects,
                'total_participants': total_participants,
                'completed': completed_projects,
                'pending': pending_projects
            },
            'projects': projects_payload,
            'recent_projects': recent_projects_payload
        })
    
    elif user_type == 'admin':
        # Pending projects for review
        pending_projects = Project.query.filter_by(status='pending').order_by(Project.created_at.desc()).all()
        projects_payload = []
        for project in pending_projects:
            org = project.organization
            projects_payload.append({
                'id': project.id,
                'title': project.title,
                'organization_name': org.display_name or org.username if org else 'Unknown',
                'organization_email': org.email if org else None,
                'date': project.date.strftime('%Y-%m-%d') if project.date else None,
                'location': project.location,
                'max_participants': project.max_participants,
                'rating': project.rating,
                'description': project.description,
                'submitted_date': project.created_at.strftime('%Y-%m-%d') if project.created_at else None
            })
        
        # Pending volunteer records for review
        pending_records = VolunteerRecord.query.filter_by(status='pending').order_by(VolunteerRecord.completed_at.desc()).all()
        records_payload = []
        for record in pending_records:
            participant = record.user
            project = record.project
            org = project.organization if project else None
            records_payload.append({
                'id': record.id,
                'participant_name': participant.display_name or participant.username,
                'project_name': project.title if project else 'Unknown',
                'organization_name': org.display_name or org.username if org else 'Unknown',
                'hours': record.hours,
                'points': record.points,
                'completion_date': record.completed_at.strftime('%Y-%m-%d') if record.completed_at else None
            })
        
        # All users (exclude admin users)
        users = User.query.filter(User.user_type != 'admin').order_by(User.created_at.desc()).limit(100).all()
        users_payload = []
        for u in users:
            users_payload.append({
                'id': u.id,
                'username': u.username,
                'display_name': u.display_name,
                'email': u.email,
                'user_type': u.user_type,
                'is_active': u.is_active if hasattr(u, 'is_active') else True,
                'created_at': u.created_at.strftime('%Y-%m-%d') if u.created_at else None
            })
        
        return jsonify({
            'pending_projects': projects_payload,
            'pending_records': records_payload,
            'users': users_payload
        })
    
    return jsonify({'error': 'Invalid user type'}), 400

# Registrations Resource
@bp.route('/api/v1/projects/<int:project_id>/registrations', methods=['GET'])
@login_required
def api_project_registrations_list(project_id):
    """Get all registrations for a project."""
    project = Project.query.get_or_404(project_id)
    
    # Check permissions
    if current_user.user_type == 'organization' and project.organization_id != current_user.id:
        return jsonify({'error': 'Unauthorized'}), 403
    elif current_user.user_type == 'participant':
        # Participants can only see their own registrations
        registrations = Registration.query.filter_by(
            project_id=project_id,
            user_id=current_user.id
        ).all()
    else:
        # Admin or organization owner can see all
        registrations = Registration.query.filter_by(project_id=project_id).all()
    
    result = []
    for reg in registrations:
        participant = reg.user
        result.append({
            'id': reg.id,
            'user_id': reg.user_id,
            'project_id': reg.project_id,
            'status': reg.status,
            'created_at': reg.created_at.strftime('%Y-%m-%d %H:%M:%S') if reg.created_at else None,
            'participant': {
                'id': participant.id,
                'name': participant.display_name or participant.username,
                'email': participant.email
            } if participant else None
        })
    
    return jsonify(result)

@bp.route('/api/v1/projects/<int:project_id>/registrations', methods=['POST'])
@login_required
def api_project_registrations_create(project_id):
    """Register for a project."""
    # Admin users are not allowed to register for projects
    if current_user.user_type == 'admin':
        return jsonify({'error': 'Admin users cannot register for projects', 'requires_login': True}), 403
    
    project = Project.query.get_or_404(project_id)
    
    # Check if already registered
    existing = Registration.query.filter_by(
        user_id=current_user.id,
        project_id=project_id
    ).first()
    
    if existing:
        return jsonify({'error': 'Already registered for this project'}), 400
    
    # Check if project is full
    current_registrations = Registration.query.filter(
        Registration.project_id == project_id,
        Registration.status.in_(('registered', 'approved'))
    ).count()
    
    if current_registrations >= project.max_participants:
        return jsonify({'error': 'Project is full'}), 400
    
    registration = Registration(
        user_id=current_user.id,
        project_id=project_id,
        status='registered'
    )
    db.session.add(registration)
    db.session.commit()
    
    return jsonify({
        'id': registration.id,
        'project_id': project_id,
        'status': registration.status,
        'message': 'Successfully registered for project'
    }), 201

@bp.route('/api/v1/registrations/<int:registration_id>', methods=['GET'])
@login_required
def api_registration_detail(registration_id):
    """Get a single registration."""
    registration = Registration.query.get_or_404(registration_id)
    
    # Check permissions
    if current_user.user_type == 'participant' and registration.user_id != current_user.id:
        return jsonify({'error': 'Unauthorized'}), 403
    elif current_user.user_type == 'organization':
        project = registration.project
        if not project or project.organization_id != current_user.id:
            return jsonify({'error': 'Unauthorized'}), 403
    
    participant = registration.user
    project = registration.project
    
    return jsonify({
        'id': registration.id,
        'user_id': registration.user_id,
        'project_id': registration.project_id,
        'status': registration.status,
        'created_at': registration.created_at.strftime('%Y-%m-%d %H:%M:%S') if registration.created_at else None,
        'participant': {
            'id': participant.id,
            'name': participant.display_name or participant.username,
            'email': participant.email
        } if participant else None,
        'project': {
            'id': project.id,
            'title': project.title
        } if project else None
    })

@bp.route('/api/v1/registrations/<int:registration_id>', methods=['PATCH'])
@login_required
def api_registration_update(registration_id):
    """Update registration status."""
    registration = Registration.query.get_or_404(registration_id)
    project = registration.project
    
    # Check permissions
    if current_user.user_type == 'organization':
        if not project or project.organization_id != current_user.id:
            return jsonify({'error': 'Unauthorized'}), 403
    elif current_user.user_type != 'admin':
        return jsonify({'error': 'Unauthorized'}), 403
    
    data = request.get_json(silent=True) or request.form or {}
    new_status = (data.get('status') or '').strip().lower()
    allowed_statuses = {'registered', 'approved', 'cancelled', 'completed'}
    
    if new_status not in allowed_statuses:
        return jsonify({'error': 'Invalid status'}), 400
    
    registration.status = new_status
    
    # If organization confirms participant completed project, auto-create pending volunteer record
    if new_status == 'completed':
        existing_record = VolunteerRecord.query.filter_by(
            user_id=registration.user_id,
            project_id=registration.project_id
        ).first()
        
        if not existing_record:
            volunteer_record = VolunteerRecord(
                user_id=registration.user_id,
                project_id=registration.project_id,
                hours=project.duration,
                points=project.points,
                status='pending',
                completed_at=datetime.utcnow()
            )
            db.session.add(volunteer_record)
    
    db.session.commit()
    
    # Check if project should be auto-completed
    project_auto_completed = _check_and_auto_complete_project(project)
    
    response_data = {
        'id': registration.id,
        'status': new_status,
        'message': 'Registration status updated successfully'
    }
    if project_auto_completed:
        response_data['project_auto_completed'] = True
        response_data['message'] = 'Registration status updated. Project has been automatically marked as completed.'
    
    return jsonify(response_data)

@bp.route('/api/v1/registrations/<int:registration_id>', methods=['DELETE'])
@login_required
def api_registration_delete(registration_id):
    """Cancel/delete a registration."""
    registration = Registration.query.get_or_404(registration_id)
    
    # Check permissions
    if current_user.user_type == 'participant' and registration.user_id != current_user.id:
        return jsonify({'error': 'Unauthorized'}), 403
    elif current_user.user_type == 'organization':
        project = registration.project
        if not project or project.organization_id != current_user.id:
            return jsonify({'error': 'Unauthorized'}), 403
    
    # Instead of deleting, mark as cancelled
    registration.status = 'cancelled'
    db.session.commit()
    
    return jsonify({'message': 'Registration cancelled successfully'}), 200


@bp.route('/api/organization/registrations/<int:project_id>')
@login_required
def api_organization_registrations(project_id):
    if current_user.user_type != 'organization':
        return jsonify({'error': 'Not authenticated'}), 401
    
    project = Project.query.get_or_404(project_id)
    if project.organization_id != current_user.id:
        return jsonify({'error': 'Unauthorized'}), 403
    
    registrations = Registration.query.filter_by(project_id=project_id).all()
    registrations_payload = []
    for reg in registrations:
        participant = reg.user
        registrations_payload.append({
            'id': reg.id,
            'participant_name': participant.display_name or participant.username,
            'participant_email': participant.email,
            'registration_date': reg.created_at.strftime('%Y-%m-%d') if reg.created_at else None,
            'status': reg.status
        })
    
    return jsonify({
        'project_title': project.title,
        'registrations': registrations_payload
    })


@bp.route('/api/organization/all-registrations')
@login_required
def api_organization_all_registrations():
    """Get all registrations for all projects of the current organization."""
    if current_user.user_type != 'organization':
        return jsonify({'error': 'Not authenticated'}), 401
    
    # Get all projects for this organization
    projects = Project.query.filter_by(organization_id=current_user.id).all()
    
    projects_with_registrations = []
    for project in projects:
        registrations = Registration.query.filter_by(project_id=project.id).all()
        registrations_payload = []
        for reg in registrations:
            participant = reg.user
            registrations_payload.append({
                'id': reg.id,
                'participant_name': participant.display_name or participant.username,
                'participant_email': participant.email,
                'registration_date': reg.created_at.strftime('%Y-%m-%d') if reg.created_at else None,
                'status': reg.status
            })
        
        projects_with_registrations.append({
            'project_id': project.id,
            'project_title': project.title,
            'project_status': project.status,
            'registrations': registrations_payload,
            'total_registrations': len(registrations_payload)
        })
    
    return jsonify({
        'projects': projects_with_registrations
    })

def _check_and_auto_complete_project(project):
    """Check if all participants are completed or cancelled, and auto-complete the project if so."""
    if project.status == 'completed':
        return False  # Already completed
    
    # Get all registrations for this project
    all_registrations = Registration.query.filter_by(project_id=project.id).all()
    
    if not all_registrations:
        return False  # No registrations, can't complete
    
    # Check if all registrations are either 'completed' or 'cancelled'
    all_finalized = all(
        reg.status in ('completed', 'cancelled') 
        for reg in all_registrations
    )
    
    if not all_finalized:
        return False  # Not all participants are finalized
    
    # Auto-complete the project
    project.status = 'completed'
    
    # Ensure volunteer records exist for completed participants
    records_created = 0
    for registration in all_registrations:
        if registration.status == 'completed':
            # Check if volunteer record already exists
            existing_record = VolunteerRecord.query.filter_by(
                user_id=registration.user_id,
                project_id=project.id
            ).first()
            
            if not existing_record:
                # Create volunteer record for confirmed participants
                volunteer_record = VolunteerRecord(
                    user_id=registration.user_id,
                    project_id=project.id,
                    hours=project.duration,
                    points=project.points,
                    status='pending',  # Wait for admin approval
                    completed_at=datetime.utcnow()
                )
                db.session.add(volunteer_record)
                records_created += 1
    
    db.session.commit()
    return True  # Project was auto-completed




# Records Resource (VolunteerRecord)
@bp.route('/api/v1/records', methods=['GET'])
@login_required
def api_records_list():
    """Get volunteer records."""
    # Support query parameters
    status = request.args.get('status')
    user_id = request.args.get('user_id', type=int)
    
    query = VolunteerRecord.query
    
    # Permission checks
    if current_user.user_type == 'participant':
        # Participants can only see their own records
        query = query.filter_by(user_id=current_user.id)
    elif current_user.user_type == 'organization':
        # Organizations can see records for their projects
        query = query.join(Project).filter(Project.organization_id == current_user.id)
    elif current_user.user_type != 'admin':
        return jsonify({'error': 'Unauthorized'}), 403
    
    if status:
        query = query.filter_by(status=status)
    if user_id and current_user.user_type == 'admin':
        query = query.filter_by(user_id=user_id)
    
    records = query.order_by(VolunteerRecord.completed_at.desc()).all()
    
    result = []
    for record in records:
        project = record.project
        participant = record.user
        organization = project.organization if project else None
        
        result.append({
            'id': record.id,
            'user_id': record.user_id,
            'project_id': record.project_id,
            'hours': record.hours,
            'points': record.points,
            'status': record.status,
            'completed_at': record.completed_at.strftime('%Y-%m-%d') if record.completed_at else None,
            'project': {
                'id': project.id,
                'title': project.title,
                'category': project.category
            } if project else None,
            'participant': {
                'id': participant.id,
                'name': participant.display_name or participant.username
            } if participant else None,
            'organization': {
                'id': organization.id,
                'name': organization.display_name or organization.username
            } if organization else None
        })
    
    return jsonify(result)

@bp.route('/api/v1/records/<int:record_id>', methods=['GET'])
@login_required
def api_record_detail(record_id):
    """Get a single volunteer record."""
    record = VolunteerRecord.query.get_or_404(record_id)
    
    # Check permissions
    if current_user.user_type == 'participant' and record.user_id != current_user.id:
        return jsonify({'error': 'Unauthorized'}), 403
    elif current_user.user_type == 'organization':
        project = record.project
        if not project or project.organization_id != current_user.id:
            return jsonify({'error': 'Unauthorized'}), 403
    
    project = record.project
    participant = record.user
    organization = project.organization if project else None
    
    return jsonify({
        'id': record.id,
        'user_id': record.user_id,
        'project_id': record.project_id,
        'hours': record.hours,
        'points': record.points,
        'status': record.status,
        'completed_at': record.completed_at.strftime('%Y-%m-%d') if record.completed_at else None,
        'project': {
            'id': project.id,
            'title': project.title,
            'category': project.category
        } if project else None,
        'participant': {
            'id': participant.id,
            'name': participant.display_name or participant.username,
            'email': participant.email
        } if participant else None,
        'organization': {
            'id': organization.id,
            'name': organization.display_name or organization.username
        } if organization else None
    })

@bp.route('/api/v1/records/<int:record_id>', methods=['PATCH'])
@login_required
def api_record_update(record_id):
    """Update a volunteer record (status)."""
    if current_user.user_type != 'admin':
        return jsonify({'error': 'Only admins can update records'}), 403
    
    record = VolunteerRecord.query.get_or_404(record_id)
    data = request.get_json() or {}
    
    if 'status' in data:
        allowed_statuses = {'pending', 'approved', 'rejected'}
        if data['status'] not in allowed_statuses:
            return jsonify({'error': 'Invalid status'}), 400
        record.status = data['status']
    
    db.session.commit()
    return jsonify({
        'id': record.id,
        'status': record.status,
        'message': 'Record updated successfully'
    })

@bp.route('/api/v1/records/batch', methods=['PATCH'])
@login_required
def api_records_batch_update():
    """Batch update volunteer records."""
    if current_user.user_type != 'admin':
        return jsonify({'error': 'Only admins can batch update records'}), 403
    
    data = request.get_json() or {}
    record_ids = data.get('record_ids', [])
    new_status = data.get('status')
    
    if not record_ids:
        return jsonify({'error': 'No record IDs provided'}), 400
    
    if not isinstance(record_ids, list):
        return jsonify({'error': 'record_ids must be a list'}), 400
    
    if new_status not in ('pending', 'approved', 'rejected'):
        return jsonify({'error': 'Invalid status'}), 400
    
    # Get all records that match the IDs and are pending (for approve action)
    if new_status == 'approved':
        records = VolunteerRecord.query.filter(
            VolunteerRecord.id.in_(record_ids),
            VolunteerRecord.status == 'pending'
        ).all()
    else:
        records = VolunteerRecord.query.filter(VolunteerRecord.id.in_(record_ids)).all()
    
    if not records:
        return jsonify({'error': 'No records found'}), 404
    
    # Update all records
    updated_count = 0
    for record in records:
        record.status = new_status
        updated_count += 1
    
    db.session.commit()
    
    return jsonify({
        'updated_count': updated_count,
        'total_requested': len(record_ids),
        'status': new_status,
        'message': f'Successfully updated {updated_count} record(s)'
    })


# Users Resource
@bp.route('/api/v1/users', methods=['GET'])
@login_required
def api_users_list():
    """Get users list (admin only, excludes admin users)."""
    if current_user.user_type != 'admin':
        return jsonify({'error': 'Unauthorized'}), 403
    
    # Exclude admin users
    users = User.query.filter(User.user_type != 'admin').order_by(User.created_at.desc()).limit(100).all()
    
    result = []
    for u in users:
        result.append({
            'id': u.id,
            'username': u.username,
            'display_name': u.display_name,
            'email': u.email,
            'user_type': u.user_type,
            'is_active': u.is_active if hasattr(u, 'is_active') else True,
            'created_at': u.created_at.strftime('%Y-%m-%d') if u.created_at else None
        })
    
    return jsonify(result)

@bp.route('/api/v1/users/me', methods=['GET'])
@login_required
def api_users_me():
    """Get current user information."""
    return jsonify({
        'id': current_user.id,
        'username': current_user.username,
        'display_name': current_user.display_name,
        'email': current_user.email,
        'user_type': current_user.user_type,
        'is_active': current_user.is_active if hasattr(current_user, 'is_active') else True,
        'created_at': current_user.created_at.strftime('%Y-%m-%d') if current_user.created_at else None
    })

@bp.route('/api/v1/users/<int:user_id>', methods=['GET'])
@login_required
def api_user_detail(user_id):
    """Get a single user."""
    user = User.query.get_or_404(user_id)
    
    # Check permissions
    if current_user.user_type == 'participant' and current_user.id != user_id:
        return jsonify({'error': 'Unauthorized'}), 403
    elif current_user.user_type == 'organization' and current_user.id != user_id:
        return jsonify({'error': 'Unauthorized'}), 403
    
    return jsonify({
        'id': user.id,
        'username': user.username,
        'display_name': user.display_name,
        'email': user.email,
        'user_type': user.user_type,
        'is_active': user.is_active if hasattr(user, 'is_active') else True,
        'created_at': user.created_at.strftime('%Y-%m-%d') if user.created_at else None
    })

@bp.route('/api/v1/users/<int:user_id>', methods=['PATCH'])
@login_required
def api_user_update(user_id):
    """Update a user (admin only for status changes)."""
    user = User.query.get_or_404(user_id)
    data = request.get_json() or {}
    
    # Prevent modifying admin users
    if user.user_type == 'admin':
        return jsonify({'error': 'Cannot modify admin user'}), 403
    
    # Check permissions
    if current_user.user_type == 'admin':
        # Admin can update is_active
        if 'is_active' in data:
            if hasattr(user, 'is_active'):
                user.is_active = bool(data['is_active'])
            else:
                return jsonify({'error': 'User status feature not available. Database migration required.'}), 500
    elif current_user.id == user_id:
        # Users can update their own profile (but not is_active)
        if 'is_active' in data:
            return jsonify({'error': 'Cannot modify your own status'}), 403
        # Allow other profile updates
        for key in ['display_name', 'description']:
            if key in data:
                setattr(user, key, data[key])
    else:
        return jsonify({'error': 'Unauthorized'}), 403
    
    db.session.commit()
    
    return jsonify({
        'id': user.id,
        'is_active': user.is_active if hasattr(user, 'is_active') else True,
        'message': 'User updated successfully'
    })


# Comments Resource
@bp.route('/api/v1/projects/<int:project_id>/comments', methods=['GET'])
def api_project_comments_list(project_id):
    """Get comments for a project."""
    project = Project.query.get_or_404(project_id)
    
    comments = Comment.query.filter_by(project_id=project_id).order_by(Comment.created_at.desc()).all()
    
    comments_data = []
    for comment in comments:
        comments_data.append({
            'id': comment.id,
            'user_id': comment.user_id,
            'user_name': comment.user.display_name or comment.user.username if comment.user else 'Unknown',
            'user_type': comment.user.user_type.title() if comment.user else 'Unknown',
            'comment': comment.content,
            'created_at': comment.created_at.strftime('%Y-%m-%d %H:%M') if comment.created_at else None
        })
    
    return jsonify(comments_data)

@bp.route('/api/v1/projects/<int:project_id>/comments', methods=['POST'])
@login_required
def api_project_comments_create(project_id):
    """Create a comment on a project."""
    # Admin users are not allowed to comment on projects
    if current_user.user_type == 'admin':
        return jsonify({'error': 'Admin users cannot comment on projects'}), 403
    
    project = Project.query.get_or_404(project_id)
    data = request.get_json()
    comment_text = data.get('comment', '').strip()
    
    if not comment_text:
        return jsonify({'error': 'Comment cannot be empty'}), 400
    
    # Check permissions: participants must be registered, organizations can comment on their own projects
    can_comment = False
    if current_user.user_type == 'participant':
        # Participant must be registered for the project
        registration = Registration.query.filter(
            Registration.user_id == current_user.id,
            Registration.project_id == project_id,
            Registration.status.in_(('registered', 'approved', 'completed'))
        ).first()
        can_comment = registration is not None
    elif current_user.user_type == 'organization':
        # Organization can comment on their own projects
        can_comment = (project.organization_id == current_user.id)
    
    if not can_comment:
        if current_user.user_type == 'participant':
            return jsonify({'error': 'You must be registered for this project to comment'}), 403
        else:
            return jsonify({'error': 'You can only comment on your own projects'}), 403
    
    # Create comment
    comment = Comment(
        project_id=project_id,
        user_id=current_user.id,
        content=comment_text
    )
    db.session.add(comment)
    db.session.commit()
    
    return jsonify({
        'id': comment.id,
        'project_id': project_id,
        'user_name': current_user.display_name or current_user.username,
        'user_type': current_user.user_type.title(),
        'comment': comment.content,
        'created_at': comment.created_at.strftime('%Y-%m-%d %H:%M') if comment.created_at else None,
        'message': 'Comment posted successfully'
    }), 201


# System Settings API
@bp.route('/api/v1/admin/settings', methods=['GET'])
@login_required
def api_get_settings():
    """Get system settings."""
    if current_user.user_type != 'admin':
        return jsonify({'error': 'Unauthorized'}), 403
    
    settings = {
        'points_per_hour': SystemSettings.get_setting('points_per_hour', '20'),
        'auto_approve_under_hours': SystemSettings.get_setting('auto_approve_under_hours', 'false'),
        'project_requires_review': SystemSettings.get_setting('project_requires_review', 'true')
    }
    
    return jsonify(settings)

@bp.route('/api/v1/admin/settings', methods=['POST'])
@login_required
def api_save_settings():
    """Save system settings."""
    if current_user.user_type != 'admin':
        return jsonify({'error': 'Unauthorized'}), 403
    
    data = request.get_json() or {}
    
    # Save points per hour
    if 'points_per_hour' in data:
        points = int(data['points_per_hour'])
        if points < 1:
            return jsonify({'error': 'Points per hour must be at least 1'}), 400
        SystemSettings.set_setting('points_per_hour', points)
    
    # Save auto-approve setting
    if 'auto_approve_under_hours' in data:
        SystemSettings.set_setting('auto_approve_under_hours', str(data['auto_approve_under_hours']).lower())
    
    # Save project review requirement
    if 'project_requires_review' in data:
        SystemSettings.set_setting('project_requires_review', str(data['project_requires_review']).lower())
    
    return jsonify({
        'message': 'Settings saved successfully',
        'settings': {
            'points_per_hour': SystemSettings.get_setting('points_per_hour', '20'),
            'auto_approve_under_hours': SystemSettings.get_setting('auto_approve_under_hours', 'false'),
            'project_requires_review': SystemSettings.get_setting('project_requires_review', 'true')
        }
    })

# Development-only: list users to help debug login issues
@bp.route('/dev/users')
def dev_users():
    if not current_app.debug and current_app.config.get('ENV') != 'development':
        return jsonify({'error': 'Not available'}), 403
    q = User.query
    user_type = request.args.get('user_type')
    if user_type:
        q = q.filter(User.user_type == user_type.strip().lower())
    users = q.order_by(User.id.desc()).limit(100).all()
    return jsonify([
        {
            'id': u.id,
            'username': u.username,
            'email': u.email,
            'user_type': u.user_type,
            'created_at': u.created_at.isoformat() if u.created_at else None
        } for u in users
    ])